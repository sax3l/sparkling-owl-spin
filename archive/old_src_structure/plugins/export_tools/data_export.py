#!/usr/bin/env python3
"""
Data Export Adapter för Sparkling-Owl-Spin
Comprehensive data export capabilities för multiple formats and destinations
"""

import logging
import asyncio
import aiofiles
import json
import csv
import io
import gzip
from typing import Dict, List, Any, Optional, Union, Iterator
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import Enum
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import xml.etree.ElementTree as ET
import yaml

logger = logging.getLogger(__name__)

class ExportFormat(Enum):
    """Supported export formats"""
    JSON = "json"
    CSV = "csv" 
    EXCEL = "xlsx"
    XML = "xml"
    YAML = "yaml"
    TXT = "txt"
    HTML = "html"
    MARKDOWN = "md"
    PARQUET = "parquet"
    SQL = "sql"

class CompressionType(Enum):
    """Compression types"""
    NONE = "none"
    GZIP = "gzip"
    ZIP = "zip"
    BZIP2 = "bz2"

class ExportDestination(Enum):
    """Export destinations"""
    LOCAL_FILE = "local_file"
    CLOUD_STORAGE = "cloud_storage"  # AWS S3, Azure Blob, GCS
    DATABASE = "database"
    API_ENDPOINT = "api_endpoint"
    EMAIL = "email"
    FTP = "ftp"

@dataclass
class ExportConfig:
    """Export configuration"""
    format: ExportFormat
    destination: ExportDestination = ExportDestination.LOCAL_FILE
    file_path: Optional[str] = None
    compression: CompressionType = CompressionType.NONE
    include_metadata: bool = True
    include_headers: bool = True
    batch_size: int = 1000
    encoding: str = "utf-8"
    delimiter: str = ","  # För CSV
    date_format: str = "%Y-%m-%d %H:%M:%S"
    pretty_print: bool = True  # För JSON/XML
    custom_fields: Optional[List[str]] = None
    filters: Optional[Dict[str, Any]] = None

@dataclass
class ExportJob:
    """Export job tracking"""
    job_id: str
    config: ExportConfig
    data_source: str
    total_records: int = 0
    processed_records: int = 0
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None
    status: str = "pending"  # pending, running, completed, failed
    error_message: Optional[str] = None
    output_file: Optional[str] = None
    file_size: Optional[int] = None
    
    @property
    def progress_percentage(self) -> float:
        """Calculate progress percentage"""
        if self.total_records == 0:
            return 0.0
        return (self.processed_records / self.total_records) * 100
        
    @property
    def duration(self) -> Optional[timedelta]:
        """Calculate job duration"""
        if self.end_time:
            return self.end_time - self.start_time
        return datetime.now() - self.start_time

class DataExportAdapter:
    """Data Export integration för comprehensive export capabilities"""
    
    def __init__(self, plugin_info):
        self.plugin_info = plugin_info
        self.initialized = False
        self.export_jobs: Dict[str, ExportJob] = {}
        self.export_dir = Path("exports")
        self.temp_dir = Path("temp_exports")
        
        # Statistics
        self.stats = {
            "total_exports": 0,
            "successful_exports": 0,
            "failed_exports": 0,
            "by_format": {},
            "by_destination": {},
            "total_records_exported": 0,
            "total_file_size": 0
        }
        
        # Format handlers
        self.format_handlers = {
            ExportFormat.JSON: self._export_json,
            ExportFormat.CSV: self._export_csv,
            ExportFormat.EXCEL: self._export_excel,
            ExportFormat.XML: self._export_xml,
            ExportFormat.YAML: self._export_yaml,
            ExportFormat.TXT: self._export_txt,
            ExportFormat.HTML: self._export_html,
            ExportFormat.MARKDOWN: self._export_markdown,
            ExportFormat.SQL: self._export_sql
        }
        
    async def initialize(self):
        """Initialize Data Export adapter"""
        try:
            logger.info("📤 Initializing Data Export Adapter")
            
            # Create directories
            self.export_dir.mkdir(exist_ok=True)
            self.temp_dir.mkdir(exist_ok=True)
            
            # Initialize statistics
            for format_type in ExportFormat:
                self.stats["by_format"][format_type.value] = {
                    "count": 0,
                    "total_size": 0
                }
                
            for dest_type in ExportDestination:
                self.stats["by_destination"][dest_type.value] = 0
            
            # Test export capabilities
            await self._test_export_capabilities()
            
            self.initialized = True
            logger.info("✅ Data Export Adapter initialized")
            
        except Exception as e:
            logger.error(f"❌ Failed to initialize Data Export: {str(e)}")
            self.initialized = True  # Continue with limited functionality
            
    async def _test_export_capabilities(self):
        """Test export capabilities"""
        logger.info("🧪 Testing export capabilities...")
        
        # Test sample data export
        test_data = [
            {"id": 1, "name": "Test Item 1", "value": 100, "date": datetime.now()},
            {"id": 2, "name": "Test Item 2", "value": 200, "date": datetime.now()}
        ]
        
        try:
            # Test JSON export
            test_config = ExportConfig(
                format=ExportFormat.JSON,
                file_path=str(self.temp_dir / "test_export.json")
            )
            
            await self._export_json(test_data, test_config)
            logger.info("✅ JSON export test passed")
            
            # Cleanup test file
            test_file = Path(test_config.file_path)
            if test_file.exists():
                test_file.unlink()
                
        except Exception as e:
            logger.warning(f"⚠️ Export capability test failed: {str(e)}")
            
    async def create_export_job(self, data: Union[List[Dict], Iterator], 
                              config: ExportConfig,
                              job_id: Optional[str] = None) -> ExportJob:
        """Create and start export job"""
        
        if not self.initialized:
            await self.initialize()
            
        # Generate job ID if not provided
        if not job_id:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            job_id = f"export_{timestamp}_{len(self.export_jobs) + 1}"
            
        # Determine total records
        if isinstance(data, list):
            total_records = len(data)
        else:
            total_records = 0  # Unknown for iterators
            
        # Create job
        job = ExportJob(
            job_id=job_id,
            config=config,
            data_source="unknown",
            total_records=total_records
        )
        
        self.export_jobs[job_id] = job
        
        # Start export in background
        asyncio.create_task(self._execute_export_job(job, data))
        
        logger.info(f"📤 Created export job: {job_id}")
        return job
        
    async def _execute_export_job(self, job: ExportJob, data: Union[List[Dict], Iterator]):
        """Execute export job"""
        try:
            job.status = "running"
            logger.info(f"🚀 Starting export job: {job.job_id}")
            
            # Determine output file path
            if not job.config.file_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}.{job.config.format.value}"
                job.config.file_path = str(self.export_dir / filename)
                
            job.output_file = job.config.file_path
            
            # Apply filters if specified
            if job.config.filters:
                data = self._apply_filters(data, job.config.filters)
                
            # Select custom fields if specified
            if job.config.custom_fields:
                data = self._select_fields(data, job.config.custom_fields)
                
            # Get format handler
            handler = self.format_handlers.get(job.config.format)
            if not handler:
                raise ValueError(f"Unsupported format: {job.config.format}")
                
            # Execute export
            await handler(data, job.config, job)
            
            # Apply compression if specified
            if job.config.compression != CompressionType.NONE:
                await self._compress_file(job)
                
            # Update job status
            job.status = "completed"
            job.end_time = datetime.now()
            
            # Update statistics
            self.stats["total_exports"] += 1
            self.stats["successful_exports"] += 1
            self.stats["by_format"][job.config.format.value]["count"] += 1
            self.stats["by_destination"][job.config.destination.value] += 1
            self.stats["total_records_exported"] += job.processed_records
            
            # Get file size
            if Path(job.output_file).exists():
                job.file_size = Path(job.output_file).stat().st_size
                self.stats["by_format"][job.config.format.value]["total_size"] += job.file_size
                self.stats["total_file_size"] += job.file_size
                
            logger.info(f"✅ Export job completed: {job.job_id} ({job.processed_records} records)")
            
        except Exception as e:
            job.status = "failed"
            job.end_time = datetime.now()
            job.error_message = str(e)
            self.stats["total_exports"] += 1
            self.stats["failed_exports"] += 1
            logger.error(f"❌ Export job failed: {job.job_id} - {str(e)}")
            
    async def _export_json(self, data: Union[List[Dict], Iterator], 
                         config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to JSON format"""
        
        async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
            if isinstance(data, list):
                # Export all data at once
                json_data = self._prepare_data_for_export(data, config)
                json_str = json.dumps(
                    json_data,
                    indent=2 if config.pretty_print else None,
                    default=self._json_serializer,
                    ensure_ascii=False
                )
                await f.write(json_str)
                
                if job:
                    job.processed_records = len(data)
                    
            else:
                # Stream data för large datasets
                await f.write('[\n')
                first_item = True
                processed = 0
                
                for item in data:
                    if not first_item:
                        await f.write(',\n')
                    else:
                        first_item = False
                        
                    prepared_item = self._prepare_item_for_export(item, config)
                    json_str = json.dumps(
                        prepared_item,
                        indent=2 if config.pretty_print else None,
                        default=self._json_serializer,
                        ensure_ascii=False
                    )
                    await f.write(json_str)
                    
                    processed += 1
                    if job:
                        job.processed_records = processed
                        
                    # Process in batches för memory efficiency
                    if processed % config.batch_size == 0:
                        await asyncio.sleep(0)  # Yield control
                        
                await f.write('\n]')
                
    async def _export_csv(self, data: Union[List[Dict], Iterator],
                        config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to CSV format"""
        
        async with aiofiles.open(config.file_path, 'w', encoding=config.encoding, newline='') as f:
            # Prepare writer
            output = io.StringIO()
            
            if isinstance(data, list) and data:
                # Get fieldnames från first item
                first_item = self._prepare_item_for_export(data[0], config)
                fieldnames = list(first_item.keys())
                
                writer = csv.DictWriter(
                    output,
                    fieldnames=fieldnames,
                    delimiter=config.delimiter,
                    quoting=csv.QUOTE_MINIMAL
                )
                
                if config.include_headers:
                    writer.writeheader()
                    
                # Write data
                prepared_data = [self._prepare_item_for_export(item, config) for item in data]
                writer.writerows(prepared_data)
                
                await f.write(output.getvalue())
                
                if job:
                    job.processed_records = len(data)
                    
            else:
                # Stream data för large datasets  
                writer = None
                processed = 0
                
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    
                    if writer is None:
                        # Initialize writer with first item
                        fieldnames = list(prepared_item.keys())
                        writer = csv.DictWriter(
                            output,
                            fieldnames=fieldnames,
                            delimiter=config.delimiter,
                            quoting=csv.QUOTE_MINIMAL
                        )
                        
                        if config.include_headers:
                            writer.writeheader()
                            await f.write(output.getvalue())
                            output = io.StringIO()
                            
                    writer.writerow(prepared_item)
                    processed += 1
                    
                    if job:
                        job.processed_records = processed
                        
                    # Write batch
                    if processed % config.batch_size == 0:
                        await f.write(output.getvalue())
                        output = io.StringIO()
                        writer = csv.DictWriter(
                            output,
                            fieldnames=fieldnames,
                            delimiter=config.delimiter,
                            quoting=csv.QUOTE_MINIMAL
                        )
                        await asyncio.sleep(0)  # Yield control
                        
                # Write remaining data
                if output.getvalue():
                    await f.write(output.getvalue())
                    
    async def _export_excel(self, data: Union[List[Dict], Iterator],
                          config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to Excel format"""
        
        # Convert data to DataFrame
        if isinstance(data, list):
            prepared_data = [self._prepare_item_for_export(item, config) for item in data]
            df = pd.DataFrame(prepared_data)
        else:
            # För iterator, collect in batches
            all_data = []
            processed = 0
            
            for item in data:
                prepared_item = self._prepare_item_for_export(item, config)
                all_data.append(prepared_item)
                processed += 1
                
                if job:
                    job.processed_records = processed
                    
                if processed % config.batch_size == 0:
                    await asyncio.sleep(0)  # Yield control
                    
            df = pd.DataFrame(all_data)
            
        # Create Excel writer
        with pd.ExcelWriter(config.file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Style the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # Header styling
            if config.include_headers:
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_name].width = adjusted_width
                
        if job and isinstance(data, list):
            job.processed_records = len(data)
            
    async def _export_xml(self, data: Union[List[Dict], Iterator],
                        config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to XML format"""
        
        root = ET.Element('data')
        
        if isinstance(data, list):
            prepared_data = [self._prepare_item_for_export(item, config) for item in data]
            
            for item in prepared_data:
                item_element = ET.SubElement(root, 'record')
                self._dict_to_xml(item, item_element)
                
            if job:
                job.processed_records = len(data)
                
        else:
            processed = 0
            for item in data:
                prepared_item = self._prepare_item_for_export(item, config)
                item_element = ET.SubElement(root, 'record')
                self._dict_to_xml(prepared_item, item_element)
                
                processed += 1
                if job:
                    job.processed_records = processed
                    
                if processed % config.batch_size == 0:
                    await asyncio.sleep(0)  # Yield control
                    
        # Write XML
        tree = ET.ElementTree(root)
        if config.pretty_print:
            self._indent_xml(root)
            
        tree.write(config.file_path, encoding=config.encoding, xml_declaration=True)
        
    def _dict_to_xml(self, data: Dict, parent: ET.Element):
        """Convert dictionary to XML elements"""
        for key, value in data.items():
            if isinstance(value, dict):
                child = ET.SubElement(parent, key)
                self._dict_to_xml(value, child)
            elif isinstance(value, list):
                for item in value:
                    child = ET.SubElement(parent, key)
                    if isinstance(item, dict):
                        self._dict_to_xml(item, child)
                    else:
                        child.text = str(item)
            else:
                child = ET.SubElement(parent, key)
                child.text = str(value) if value is not None else ""
                
    def _indent_xml(self, elem, level=0):
        """Add indentation to XML för pretty printing"""
        indent = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = indent + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = indent
            for elem in elem:
                self._indent_xml(elem, level + 1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = indent
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = indent
                
    async def _export_yaml(self, data: Union[List[Dict], Iterator],
                         config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to YAML format"""
        
        if isinstance(data, list):
            prepared_data = [self._prepare_item_for_export(item, config) for item in data]
            
            async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
                yaml_str = yaml.dump(
                    prepared_data,
                    default_flow_style=False,
                    allow_unicode=True,
                    sort_keys=False
                )
                await f.write(yaml_str)
                
            if job:
                job.processed_records = len(data)
        else:
            # För iterators, collect all data first
            all_data = []
            processed = 0
            
            for item in data:
                prepared_item = self._prepare_item_for_export(item, config)
                all_data.append(prepared_item)
                processed += 1
                
                if job:
                    job.processed_records = processed
                    
                if processed % config.batch_size == 0:
                    await asyncio.sleep(0)  # Yield control
                    
            async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
                yaml_str = yaml.dump(
                    all_data,
                    default_flow_style=False,
                    allow_unicode=True,
                    sort_keys=False
                )
                await f.write(yaml_str)
                
    async def _export_txt(self, data: Union[List[Dict], Iterator],
                        config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to plain text format"""
        
        async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
            if isinstance(data, list):
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    line = self._format_item_as_text(prepared_item)
                    await f.write(line + '\n')
                    
                if job:
                    job.processed_records = len(data)
            else:
                processed = 0
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    line = self._format_item_as_text(prepared_item)
                    await f.write(line + '\n')
                    
                    processed += 1
                    if job:
                        job.processed_records = processed
                        
                    if processed % config.batch_size == 0:
                        await asyncio.sleep(0)  # Yield control
                        
    def _format_item_as_text(self, item: Dict) -> str:
        """Format dictionary as readable text"""
        lines = []
        for key, value in item.items():
            lines.append(f"{key}: {value}")
        return " | ".join(lines)
        
    async def _export_html(self, data: Union[List[Dict], Iterator],
                         config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to HTML format"""
        
        if isinstance(data, list):
            prepared_data = [self._prepare_item_for_export(item, config) for item in data]
            df = pd.DataFrame(prepared_data)
        else:
            all_data = []
            processed = 0
            
            for item in data:
                prepared_item = self._prepare_item_for_export(item, config)
                all_data.append(prepared_item)
                processed += 1
                
                if job:
                    job.processed_records = processed
                    
                if processed % config.batch_size == 0:
                    await asyncio.sleep(0)  # Yield control
                    
            df = pd.DataFrame(all_data)
            
        # Generate HTML
        html_str = df.to_html(index=False, escape=False, table_id="data-table")
        
        # Add CSS styling
        css = """
        <style>
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #f2f2f2; font-weight: bold; }
            tr:nth-child(even) { background-color: #f9f9f9; }
        </style>
        """
        
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Data Export</title>
            {css}
        </head>
        <body>
            <h1>Data Export</h1>
            <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            {html_str}
        </body>
        </html>
        """
        
        async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
            await f.write(full_html)
            
        if job and isinstance(data, list):
            job.processed_records = len(data)
            
    async def _export_markdown(self, data: Union[List[Dict], Iterator],
                             config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data to Markdown format"""
        
        async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
            await f.write("# Data Export\n\n")
            await f.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            if isinstance(data, list) and data:
                # Create table header
                first_item = self._prepare_item_for_export(data[0], config)
                headers = list(first_item.keys())
                
                header_line = "| " + " | ".join(headers) + " |\n"
                separator_line = "| " + " | ".join(["---"] * len(headers)) + " |\n"
                
                await f.write(header_line)
                await f.write(separator_line)
                
                # Write data rows
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    row_values = [str(prepared_item.get(header, "")) for header in headers]
                    row_line = "| " + " | ".join(row_values) + " |\n"
                    await f.write(row_line)
                    
                if job:
                    job.processed_records = len(data)
            else:
                # Stream data
                headers = None
                processed = 0
                
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    
                    if headers is None:
                        headers = list(prepared_item.keys())
                        header_line = "| " + " | ".join(headers) + " |\n"
                        separator_line = "| " + " | ".join(["---"] * len(headers)) + " |\n"
                        await f.write(header_line)
                        await f.write(separator_line)
                        
                    row_values = [str(prepared_item.get(header, "")) for header in headers]
                    row_line = "| " + " | ".join(row_values) + " |\n"
                    await f.write(row_line)
                    
                    processed += 1
                    if job:
                        job.processed_records = processed
                        
                    if processed % config.batch_size == 0:
                        await asyncio.sleep(0)  # Yield control
                        
    async def _export_sql(self, data: Union[List[Dict], Iterator],
                        config: ExportConfig, job: Optional[ExportJob] = None):
        """Export data as SQL INSERT statements"""
        
        table_name = "exported_data"
        
        async with aiofiles.open(config.file_path, 'w', encoding=config.encoding) as f:
            await f.write(f"-- Data Export SQL\n")
            await f.write(f"-- Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            if isinstance(data, list) and data:
                first_item = self._prepare_item_for_export(data[0], config)
                columns = list(first_item.keys())
                
                # Create table statement
                await f.write(f"CREATE TABLE {table_name} (\n")
                column_defs = []
                for col in columns:
                    column_defs.append(f"  {col} TEXT")
                await f.write(",\n".join(column_defs))
                await f.write("\n);\n\n")
                
                # Insert statements
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    values = []
                    for col in columns:
                        value = prepared_item.get(col, "")
                        if value is None:
                            values.append("NULL")
                        else:
                            escaped_value = str(value).replace("'", "''")
                            values.append(f"'{escaped_value}'")
                    
                    values_str = ", ".join(values)
                    insert_stmt = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({values_str});\n"
                    await f.write(insert_stmt)
                    
                if job:
                    job.processed_records = len(data)
            else:
                # Stream data
                columns = None
                processed = 0
                
                for item in data:
                    prepared_item = self._prepare_item_for_export(item, config)
                    
                    if columns is None:
                        columns = list(prepared_item.keys())
                        await f.write(f"CREATE TABLE {table_name} (\n")
                        column_defs = []
                        for col in columns:
                            column_defs.append(f"  {col} TEXT")
                        await f.write(",\n".join(column_defs))
                        await f.write("\n);\n\n")
                        
                    values = []
                    for col in columns:
                        value = prepared_item.get(col, "")
                        if value is None:
                            values.append("NULL")
                        else:
                            escaped_value = str(value).replace("'", "''")
                            values.append(f"'{escaped_value}'")
                    
                    values_str = ", ".join(values)
                    insert_stmt = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({values_str});\n"
                    await f.write(insert_stmt)
                    
                    processed += 1
                    if job:
                        job.processed_records = processed
                        
                    if processed % config.batch_size == 0:
                        await asyncio.sleep(0)  # Yield control
                        
    def _apply_filters(self, data: Union[List[Dict], Iterator], 
                      filters: Dict[str, Any]) -> Union[List[Dict], Iterator]:
        """Apply filters to data"""
        
        def item_matches_filters(item: Dict) -> bool:
            for key, filter_value in filters.items():
                if key not in item:
                    return False
                    
                item_value = item[key]
                
                if isinstance(filter_value, dict):
                    # Complex filter operations
                    if '$eq' in filter_value:
                        if item_value != filter_value['$eq']:
                            return False
                    if '$ne' in filter_value:
                        if item_value == filter_value['$ne']:
                            return False
                    if '$gt' in filter_value:
                        if not (item_value > filter_value['$gt']):
                            return False
                    if '$gte' in filter_value:
                        if not (item_value >= filter_value['$gte']):
                            return False
                    if '$lt' in filter_value:
                        if not (item_value < filter_value['$lt']):
                            return False
                    if '$lte' in filter_value:
                        if not (item_value <= filter_value['$lte']):
                            return False
                    if '$in' in filter_value:
                        if item_value not in filter_value['$in']:
                            return False
                    if '$contains' in filter_value:
                        if filter_value['$contains'] not in str(item_value):
                            return False
                else:
                    # Simple equality filter
                    if item_value != filter_value:
                        return False
                        
            return True
            
        if isinstance(data, list):
            return [item for item in data if item_matches_filters(item)]
        else:
            return (item for item in data if item_matches_filters(item))
            
    def _select_fields(self, data: Union[List[Dict], Iterator],
                      fields: List[str]) -> Union[List[Dict], Iterator]:
        """Select only specified fields från data"""
        
        def select_item_fields(item: Dict) -> Dict:
            return {field: item.get(field) for field in fields if field in item}
            
        if isinstance(data, list):
            return [select_item_fields(item) for item in data]
        else:
            return (select_item_fields(item) for item in data)
            
    def _prepare_data_for_export(self, data: List[Dict], config: ExportConfig) -> List[Dict]:
        """Prepare data för export"""
        return [self._prepare_item_for_export(item, config) for item in data]
        
    def _prepare_item_for_export(self, item: Dict, config: ExportConfig) -> Dict:
        """Prepare individual item för export"""
        prepared = {}
        
        for key, value in item.items():
            # Handle datetime objects
            if isinstance(value, datetime):
                prepared[key] = value.strftime(config.date_format)
            # Handle None values
            elif value is None:
                prepared[key] = ""
            # Handle complex objects
            elif isinstance(value, (dict, list)):
                if config.format == ExportFormat.JSON:
                    prepared[key] = value
                else:
                    prepared[key] = json.dumps(value, default=self._json_serializer)
            else:
                prepared[key] = value
                
        return prepared
        
    def _json_serializer(self, obj):
        """JSON serializer för complex objects"""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif hasattr(obj, '__dict__'):
            return obj.__dict__
        else:
            return str(obj)
            
    async def _compress_file(self, job: ExportJob):
        """Compress exported file"""
        original_path = Path(job.output_file)
        
        if job.config.compression == CompressionType.GZIP:
            compressed_path = original_path.with_suffix(original_path.suffix + '.gz')
            
            with open(original_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    f_out.write(f_in.read())
                    
            # Remove original file
            original_path.unlink()
            job.output_file = str(compressed_path)
            
        # Add support för other compression types as needed
        
    def get_export_job(self, job_id: str) -> Optional[ExportJob]:
        """Get export job by ID"""
        return self.export_jobs.get(job_id)
        
    def get_all_jobs(self) -> List[ExportJob]:
        """Get all export jobs"""
        return list(self.export_jobs.values())
        
    def get_active_jobs(self) -> List[ExportJob]:
        """Get active export jobs"""
        return [job for job in self.export_jobs.values() if job.status in ["pending", "running"]]
        
    def get_export_statistics(self) -> Dict[str, Any]:
        """Get export statistics"""
        return {
            "total_exports": self.stats["total_exports"],
            "successful_exports": self.stats["successful_exports"],
            "failed_exports": self.stats["failed_exports"],
            "success_rate": (
                self.stats["successful_exports"] / max(1, self.stats["total_exports"])
            ) * 100,
            "by_format": self.stats["by_format"],
            "by_destination": self.stats["by_destination"],
            "total_records_exported": self.stats["total_records_exported"],
            "total_file_size": self.stats["total_file_size"],
            "total_jobs": len(self.export_jobs),
            "active_jobs": len(self.get_active_jobs())
        }
        
    async def cleanup_old_exports(self, older_than_days: int = 30):
        """Clean up old export files"""
        cutoff_date = datetime.now() - timedelta(days=older_than_days)
        cleaned_files = 0
        
        for export_path in [self.export_dir, self.temp_dir]:
            if export_path.exists():
                for file_path in export_path.iterdir():
                    if file_path.is_file():
                        file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                        if file_time < cutoff_date:
                            try:
                                file_path.unlink()
                                cleaned_files += 1
                            except Exception as e:
                                logger.warning(f"⚠️ Failed to delete old export file {file_path}: {str(e)}")
                                
        logger.info(f"🧹 Cleaned up {cleaned_files} old export files")
        
    async def cleanup(self):
        """Cleanup Data Export adapter"""
        logger.info("🧹 Cleaning up Data Export Adapter")
        
        # Cancel active jobs
        active_jobs = self.get_active_jobs()
        for job in active_jobs:
            job.status = "cancelled"
            job.end_time = datetime.now()
            
        self.export_jobs.clear()
        self.stats.clear()
        self.initialized = False
        logger.info("✅ Data Export Adapter cleanup completed")
