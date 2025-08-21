"""
Service layer for data export functionality.
"""

import csv
import json
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Union, BinaryIO
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from fastapi import HTTPException, status

from ..models import CrawlJob, ScrapeJob, User
from ..schemas.jobs import JobStatus


class ExportService:
    """Service for exporting scraped data in various formats."""
    
    def __init__(self, db: Session, export_dir: Optional[Path] = None):
        self.db = db
        self.export_dir = export_dir or Path("data/exports")
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def export_crawl_results(
        self,
        job_id: str,
        format: str = "json",
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Export crawl job results.
        
        Args:
            job_id: Crawl job ID
            format: Export format (json, csv, xlsx)
            user_id: User ID for permission check
            
        Returns:
            Export metadata and file info
            
        Raises:
            HTTPException: If job not found or access denied
        """
        # Get job
        job = self.db.query(CrawlJob).filter(CrawlJob.id == job_id).first()
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Crawl job not found"
            )
        
        # Check permissions
        if user_id and job.user_id != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        
        # Check if job is completed
        if job.status != JobStatus.COMPLETED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Job is not completed yet"
            )
        
        # Get results data
        results_data = job.results or []
        if not results_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No results found for this job"
            )
        
        # Generate export based on format
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"crawl_job_{job_id}_{timestamp}"
        
        if format.lower() == "json":
            return self._export_to_json(results_data, filename, job)
        elif format.lower() == "csv":
            return self._export_to_csv(results_data, filename, job)
        elif format.lower() == "xlsx":
            return self._export_to_xlsx(results_data, filename, job)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported export format: {format}"
            )
    
    def export_scrape_results(
        self,
        job_id: str,
        format: str = "json",
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Export scrape job results.
        
        Args:
            job_id: Scrape job ID
            format: Export format (json, csv, xlsx)
            user_id: User ID for permission check
            
        Returns:
            Export metadata and file info
            
        Raises:
            HTTPException: If job not found or access denied
        """
        # Get job
        job = self.db.query(ScrapeJob).filter(ScrapeJob.id == job_id).first()
        if not job:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Scrape job not found"
            )
        
        # Check permissions
        if user_id and job.user_id != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        
        # Check if job is completed
        if job.status != JobStatus.COMPLETED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Job is not completed yet"
            )
        
        # Get results data
        results_data = job.results or []
        if not results_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No results found for this job"
            )
        
        # Generate export based on format
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"scrape_job_{job_id}_{timestamp}"
        
        if format.lower() == "json":
            return self._export_to_json(results_data, filename, job)
        elif format.lower() == "csv":
            return self._export_to_csv(results_data, filename, job)
        elif format.lower() == "xlsx":
            return self._export_to_xlsx(results_data, filename, job)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported export format: {format}"
            )
    
    def get_user_exports(self, user_id: str, limit: int = 50) -> List[Dict[str, Any]]:
        """
        Get user's export history.
        
        Args:
            user_id: User ID
            limit: Maximum number of exports to return
            
        Returns:
            List of export metadata
        """
        # Get user's completed jobs
        crawl_jobs = self.db.query(CrawlJob).filter(
            CrawlJob.user_id == user_id,
            CrawlJob.status == JobStatus.COMPLETED
        ).order_by(CrawlJob.completed_at.desc()).limit(limit).all()
        
        scrape_jobs = self.db.query(ScrapeJob).filter(
            ScrapeJob.user_id == user_id,
            ScrapeJob.status == JobStatus.COMPLETED
        ).order_by(ScrapeJob.completed_at.desc()).limit(limit).all()
        
        exports = []
        
        # Add crawl jobs
        for job in crawl_jobs:
            exports.append({
                "id": str(job.id),
                "type": "crawl",
                "name": job.name or f"Crawl Job {job.id}",
                "status": job.status,
                "completed_at": job.completed_at,
                "result_count": len(job.results or []),
                "config": job.config
            })
        
        # Add scrape jobs
        for job in scrape_jobs:
            exports.append({
                "id": str(job.id),
                "type": "scrape",
                "name": job.name or f"Scrape Job {job.id}",
                "status": job.status,
                "completed_at": job.completed_at,
                "result_count": len(job.results or []),
                "urls": job.urls
            })
        
        # Sort by completion date
        exports.sort(key=lambda x: x["completed_at"] or datetime.min, reverse=True)
        
        return exports[:limit]
    
    def _export_to_json(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        job: Union[CrawlJob, ScrapeJob]
    ) -> Dict[str, Any]:
        """Export data to JSON format."""
        filepath = self.export_dir / "json" / f"{filename}.json"
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        export_data = {
            "metadata": {
                "job_id": str(job.id),
                "job_type": "crawl" if isinstance(job, CrawlJob) else "scrape",
                "exported_at": datetime.now().isoformat(),
                "total_records": len(data),
                "format": "json"
            },
            "data": data
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        return {
            "filename": filepath.name,
            "filepath": str(filepath),
            "format": "json",
            "size_bytes": filepath.stat().st_size,
            "record_count": len(data),
            "created_at": datetime.now().isoformat()
        }
    
    def _export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        job: Union[CrawlJob, ScrapeJob]
    ) -> Dict[str, Any]:
        """Export data to CSV format."""
        filepath = self.export_dir / "csv" / f"{filename}.csv"
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        if not data:
            # Create empty CSV with headers
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["No data available"])
        else:
            # Get all unique keys from all records
            all_keys = set()
            for record in data:
                if isinstance(record, dict):
                    all_keys.update(record.keys())
            
            fieldnames = sorted(list(all_keys))
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for record in data:
                    if isinstance(record, dict):
                        # Flatten nested objects
                        flattened = self._flatten_dict(record)
                        writer.writerow(flattened)
        
        return {
            "filename": filepath.name,
            "filepath": str(filepath),
            "format": "csv",
            "size_bytes": filepath.stat().st_size,
            "record_count": len(data),
            "created_at": datetime.now().isoformat()
        }
    
    def _export_to_xlsx(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        job: Union[CrawlJob, ScrapeJob]
    ) -> Dict[str, Any]:
        """Export data to Excel format."""
        filepath = self.export_dir / "excel" / f"{filename}.xlsx"
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export Data"
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if not data:
            worksheet.cell(row=1, column=1, value="No data available")
        else:
            # Get all unique keys
            all_keys = set()
            for record in data:
                if isinstance(record, dict):
                    all_keys.update(record.keys())
            
            fieldnames = sorted(list(all_keys))
            
            # Write headers
            for col, fieldname in enumerate(fieldnames, 1):
                cell = worksheet.cell(row=1, column=col, value=fieldname)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row, record in enumerate(data, 2):
                if isinstance(record, dict):
                    flattened = self._flatten_dict(record)
                    for col, fieldname in enumerate(fieldnames, 1):
                        value = flattened.get(fieldname, "")
                        # Convert complex types to strings
                        if isinstance(value, (list, dict)):
                            value = json.dumps(value, ensure_ascii=False)
                        worksheet.cell(row=row, column=col, value=str(value) if value is not None else "")
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        metadata_sheet = workbook.create_sheet("Metadata")
        metadata_sheet.cell(row=1, column=1, value="Job ID").font = header_font
        metadata_sheet.cell(row=1, column=2, value=str(job.id))
        metadata_sheet.cell(row=2, column=1, value="Job Type").font = header_font
        metadata_sheet.cell(row=2, column=2, value="crawl" if isinstance(job, CrawlJob) else "scrape")
        metadata_sheet.cell(row=3, column=1, value="Exported At").font = header_font
        metadata_sheet.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        metadata_sheet.cell(row=4, column=1, value="Total Records").font = header_font
        metadata_sheet.cell(row=4, column=2, value=len(data))
        
        workbook.save(filepath)
        
        return {
            "filename": filepath.name,
            "filepath": str(filepath),
            "format": "xlsx",
            "size_bytes": filepath.stat().st_size,
            "record_count": len(data),
            "created_at": datetime.now().isoformat()
        }
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
        """
        Flatten a nested dictionary.
        
        Args:
            d: Dictionary to flatten
            parent_key: Parent key prefix
            sep: Separator for nested keys
            
        Returns:
            Flattened dictionary
        """
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep).items())
            elif isinstance(v, list):
                # Convert lists to JSON strings for CSV/Excel
                items.append((new_key, json.dumps(v, ensure_ascii=False)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def generate_export_stream(
        self,
        data: List[Dict[str, Any]],
        format: str,
        job_metadata: Dict[str, Any]
    ) -> io.BytesIO:
        """
        Generate export data as a stream for direct download.
        
        Args:
            data: Data to export
            format: Export format
            job_metadata: Job metadata
            
        Returns:
            BytesIO stream with export data
        """
        stream = io.BytesIO()
        
        if format.lower() == "json":
            export_data = {
                "metadata": job_metadata,
                "data": data
            }
            json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
            stream.write(json_str.encode('utf-8'))
        
        elif format.lower() == "csv":
            if data:
                all_keys = set()
                for record in data:
                    if isinstance(record, dict):
                        all_keys.update(record.keys())
                
                fieldnames = sorted(list(all_keys))
                
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                
                for record in data:
                    if isinstance(record, dict):
                        flattened = self._flatten_dict(record)
                        writer.writerow(flattened)
                
                stream.write(output.getvalue().encode('utf-8'))
                output.close()
        
        stream.seek(0)
        return stream
