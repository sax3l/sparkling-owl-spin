"""Excel exporter for scraped data."""

import os
from typing import Any, Dict, List, Optional
from pathlib import Path
import logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base import BaseExporter, ExportResult, ExportConfig, ExportFormat


class ExcelExporter(BaseExporter):
    """Export data to Excel format."""
    
    def __init__(self, config: ExportConfig):
        super().__init__(config)
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        # Excel-specific options
        self.sheet_name = config.options.get('sheet_name', 'Data')
        self.max_rows_per_sheet = config.options.get('max_rows_per_sheet', 1048576)  # Excel limit
        self.auto_filter = config.options.get('auto_filter', True)
        self.freeze_header = config.options.get('freeze_header', True)
        self.style_header = config.options.get('style_header', True)
        self.auto_width = config.options.get('auto_width', True)
        self.include_formulas = config.options.get('include_formulas', False)
    
    def export(self, data: List[Dict[str, Any]], metadata: Optional[Dict[str, Any]] = None) -> ExportResult:
        """
        Export data to Excel file.
        
        Args:
            data: List of records to export
            metadata: Optional metadata about the export
            
        Returns:
            ExportResult with export details
        """
        try:
            if not data:
                return ExportResult(
                    success=True,
                    records_exported=0,
                    destination=self.config.destination,
                    format=self.config.format,
                    warnings=["No data to export"]
                )
            
            # Prepare data
            metadata = self._add_standard_metadata(metadata)
            prepared_data = self._prepare_data(data, metadata)
            
            # Ensure destination directory exists
            destination_path = Path(self.config.destination)
            destination_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            records_exported = 0
            
            # Split data into multiple sheets if necessary
            sheets_data = self._split_data_into_sheets(prepared_data)
            
            for sheet_index, (sheet_name, sheet_data) in enumerate(sheets_data.items()):
                worksheet = workbook.create_sheet(title=sheet_name)
                
                # Get field names from first record
                if sheet_data:
                    fieldnames = list(sheet_data[0].keys())
                    
                    # Write headers
                    if self.config.include_headers:
                        for col_idx, fieldname in enumerate(fieldnames, 1):
                            cell = worksheet.cell(row=1, column=col_idx, value=fieldname)
                            if self.style_header:
                                self._style_header_cell(cell)
                    
                    # Write data
                    start_row = 2 if self.config.include_headers else 1
                    
                    for row_idx, record in enumerate(sheet_data, start_row):
                        for col_idx, fieldname in enumerate(fieldnames, 1):
                            value = record.get(fieldname)
                            
                            # Convert value to Excel-compatible format
                            excel_value = self._convert_value_for_excel(value)
                            
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=excel_value)
                            self._style_data_cell(cell, value)
                        
                        records_exported += 1
                    
                    # Apply formatting
                    self._apply_worksheet_formatting(worksheet, fieldnames)
            
            # Add metadata sheet if requested
            if self.config.include_metadata and metadata:
                self._add_metadata_sheet(workbook, metadata)
            
            # Save workbook
            workbook.save(destination_path)
            
            # Get file size
            file_size = destination_path.stat().st_size if destination_path.exists() else 0
            
            self.logger.info(f"Successfully exported {records_exported} records to {self.config.destination}")
            
            return ExportResult(
                success=True,
                records_exported=records_exported,
                destination=self.config.destination,
                format=self.config.format,
                file_size=file_size
            )
            
        except Exception as e:
            return self._handle_errors(e, "Excel export")
    
    def validate_destination(self) -> bool:
        """
        Validate that the Excel destination is writable.
        
        Returns:
            True if destination is valid, False otherwise
        """
        try:
            destination_path = Path(self.config.destination)
            
            # Check file extension
            if not destination_path.suffix.lower() in ['.xlsx', '.xlsm']:
                self.logger.error(f"Invalid file extension: {destination_path.suffix}. Use .xlsx or .xlsm")
                return False
            
            # Check if parent directory exists or can be created
            parent_dir = destination_path.parent
            if not parent_dir.exists():
                try:
                    parent_dir.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    self.logger.error(f"Cannot create destination directory: {e}")
                    return False
            
            # Check if we can write to the directory
            if not os.access(parent_dir, os.W_OK):
                self.logger.error(f"No write permission for directory: {parent_dir}")
                return False
            
            # If file exists, check if we can overwrite it
            if destination_path.exists():
                if not os.access(destination_path, os.W_OK):
                    self.logger.error(f"No write permission for file: {destination_path}")
                    return False
            
            return True
            
        except Exception as e:
            self.logger.error(f"Destination validation failed: {e}")
            return False
    
    def append_data(self, data: List[Dict[str, Any]], metadata: Optional[Dict[str, Any]] = None) -> ExportResult:
        """
        Append data to existing Excel file.
        
        Args:
            data: List of records to append
            metadata: Optional metadata about the export
            
        Returns:
            ExportResult with append details
        """
        try:
            if not data:
                return ExportResult(
                    success=True,
                    records_exported=0,
                    destination=self.config.destination,
                    format=self.config.format,
                    warnings=["No data to append"]
                )
            
            destination_path = Path(self.config.destination)
            
            # If file doesn't exist, create new export
            if not destination_path.exists():
                return self.export(data, metadata)
            
            # Load existing workbook
            workbook = openpyxl.load_workbook(destination_path)
            
            # Prepare data
            metadata = self._add_standard_metadata(metadata)
            prepared_data = self._prepare_data(data, metadata)
            
            records_exported = 0
            
            # Find the data sheet (first sheet that's not metadata)
            data_sheet = None
            for sheet_name in workbook.sheetnames:
                if sheet_name != 'Metadata':
                    data_sheet = workbook[sheet_name]
                    break
            
            if not data_sheet:
                # Create new data sheet
                data_sheet = workbook.create_sheet(title=self.sheet_name)
            
            # Find last row with data
            last_row = data_sheet.max_row
            
            # Get field names from prepared data
            if prepared_data:
                fieldnames = list(prepared_data[0].keys())
                
                # Check if headers exist and match
                if last_row == 1 and self.config.include_headers:
                    # File might be empty or have only headers
                    existing_headers = []
                    for col in range(1, data_sheet.max_column + 1):
                        cell_value = data_sheet.cell(row=1, column=col).value
                        if cell_value:
                            existing_headers.append(cell_value)
                    
                    if not existing_headers:
                        # Write headers
                        for col_idx, fieldname in enumerate(fieldnames, 1):
                            cell = data_sheet.cell(row=1, column=col_idx, value=fieldname)
                            if self.style_header:
                                self._style_header_cell(cell)
                        last_row = 1
                
                # Append data starting from next row
                start_row = last_row + 1
                
                for row_idx, record in enumerate(prepared_data, start_row):
                    for col_idx, fieldname in enumerate(fieldnames, 1):
                        value = record.get(fieldname)
                        excel_value = self._convert_value_for_excel(value)
                        
                        cell = data_sheet.cell(row=row_idx, column=col_idx, value=excel_value)
                        self._style_data_cell(cell, value)
                    
                    records_exported += 1
                
                # Update formatting
                self._apply_worksheet_formatting(data_sheet, fieldnames)
            
            # Update metadata sheet
            if self.config.include_metadata and metadata:
                self._update_metadata_sheet(workbook, metadata)
            
            # Save workbook
            workbook.save(destination_path)
            
            self.logger.info(f"Successfully appended {records_exported} records to {self.config.destination}")
            
            return ExportResult(
                success=True,
                records_exported=records_exported,
                destination=self.config.destination,
                format=self.config.format,
                file_size=destination_path.stat().st_size if destination_path.exists() else 0
            )
            
        except Exception as e:
            return self._handle_errors(e, "Excel append")
    
    def read_excel_data(self, sheet_name: Optional[str] = None, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Read data from Excel file.
        
        Args:
            sheet_name: Name of sheet to read (first data sheet if None)
            limit: Maximum number of records to read
            
        Returns:
            List of records from the Excel file
        """
        try:
            destination_path = Path(self.config.destination)
            
            if not destination_path.exists():
                return []
            
            workbook = openpyxl.load_workbook(destination_path, data_only=True)
            
            # Select sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    self.logger.warning(f"Sheet '{sheet_name}' not found")
                    return []
                worksheet = workbook[sheet_name]
            else:
                # Use first sheet that's not metadata
                worksheet = None
                for name in workbook.sheetnames:
                    if name != 'Metadata':
                        worksheet = workbook[name]
                        break
                
                if not worksheet:
                    return []
            
            data = []
            
            # Get headers from first row
            headers = []
            for col in range(1, worksheet.max_column + 1):
                header = worksheet.cell(row=1, column=col).value
                if header:
                    headers.append(str(header))
                else:
                    break
            
            if not headers:
                return []
            
            # Read data rows
            start_row = 2 if self.config.include_headers else 1
            for row_idx in range(start_row, worksheet.max_row + 1):
                if limit and len(data) >= limit:
                    break
                
                record = {}
                has_data = False
                
                for col_idx, header in enumerate(headers, 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        has_data = True
                    record[header] = cell_value
                
                if has_data:
                    data.append(record)
                else:
                    break  # Stop at first empty row
            
            return data
            
        except Exception as e:
            self.logger.error(f"Failed to read Excel data: {e}")
            return []
    
    def get_file_info(self) -> Dict[str, Any]:
        """
        Get information about the Excel file.
        
        Returns:
            Dictionary with file information
        """
        try:
            destination_path = Path(self.config.destination)
            
            if not destination_path.exists():
                return {"exists": False}
            
            stat = destination_path.stat()
            
            # Get workbook info
            try:
                workbook = openpyxl.load_workbook(destination_path, data_only=True)
                
                sheets_info = []
                total_records = 0
                
                for sheet_name in workbook.sheetnames:
                    if sheet_name != 'Metadata':
                        worksheet = workbook[sheet_name]
                        row_count = worksheet.max_row
                        if self.config.include_headers and row_count > 0:
                            row_count -= 1
                        
                        sheets_info.append({
                            "name": sheet_name,
                            "rows": row_count,
                            "columns": worksheet.max_column
                        })
                        total_records += row_count
                
                return {
                    "exists": True,
                    "size_bytes": stat.st_size,
                    "total_records": total_records,
                    "sheets": sheets_info,
                    "modified_time": stat.st_mtime,
                    "created_time": stat.st_ctime,
                    "has_metadata": 'Metadata' in workbook.sheetnames
                }
                
            except Exception as e:
                return {
                    "exists": True,
                    "size_bytes": stat.st_size,
                    "modified_time": stat.st_mtime,
                    "created_time": stat.st_ctime,
                    "error": f"Could not read Excel file: {str(e)}"
                }
            
        except Exception as e:
            self.logger.error(f"Failed to get file info: {e}")
            return {"exists": False, "error": str(e)}
    
    def _split_data_into_sheets(self, data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """Split data into multiple sheets if it exceeds row limits."""
        if len(data) <= self.max_rows_per_sheet:
            return {self.sheet_name: data}
        
        sheets = {}
        for i in range(0, len(data), self.max_rows_per_sheet):
            sheet_num = (i // self.max_rows_per_sheet) + 1
            sheet_name = f"{self.sheet_name}_{sheet_num}"
            sheet_data = data[i:i + self.max_rows_per_sheet]
            sheets[sheet_name] = sheet_data
        
        return sheets
    
    def _convert_value_for_excel(self, value: Any) -> Any:
        """Convert value to Excel-compatible format."""
        if value is None:
            return ""
        elif isinstance(value, (list, dict)):
            import json
            return json.dumps(value)
        elif isinstance(value, bool):
            return value
        elif isinstance(value, (int, float)):
            return value
        elif isinstance(value, datetime):
            return value
        else:
            return str(value)
    
    def _style_header_cell(self, cell) -> None:
        """Apply styling to header cell."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def _style_data_cell(self, cell, original_value: Any) -> None:
        """Apply styling to data cell based on value type."""
        cell.border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        
        # Apply specific formatting based on data type
        if isinstance(original_value, (int, float)):
            cell.alignment = Alignment(horizontal="right")
        elif isinstance(original_value, datetime):
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")
    
    def _apply_worksheet_formatting(self, worksheet, fieldnames: List[str]) -> None:
        """Apply formatting to the entire worksheet."""
        if self.auto_width:
            # Auto-adjust column widths
            for col_idx, fieldname in enumerate(fieldnames, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(fieldname)
                
                for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                # Set column width (with some padding)
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        if self.auto_filter and self.config.include_headers:
            # Add auto filter to headers
            last_column = get_column_letter(len(fieldnames))
            worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"
        
        if self.freeze_header and self.config.include_headers:
            # Freeze header row
            worksheet.freeze_panes = "A2"
    
    def _add_metadata_sheet(self, workbook, metadata: Dict[str, Any]) -> None:
        """Add metadata sheet to workbook."""
        metadata_sheet = workbook.create_sheet(title="Metadata")
        
        # Write metadata as key-value pairs
        row = 1
        for key, value in metadata.items():
            metadata_sheet.cell(row=row, column=1, value=str(key))
            metadata_sheet.cell(row=row, column=2, value=str(value))
            row += 1
        
        # Style metadata sheet
        metadata_sheet.column_dimensions['A'].width = 20
        metadata_sheet.column_dimensions['B'].width = 40
    
    def _update_metadata_sheet(self, workbook, metadata: Dict[str, Any]) -> None:
        """Update existing metadata sheet."""
        if 'Metadata' in workbook.sheetnames:
            workbook.remove(workbook['Metadata'])
        
        self._add_metadata_sheet(workbook, metadata)
